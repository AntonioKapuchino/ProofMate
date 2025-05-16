import os
import json
import nbformat
import logging
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from typing import List, Dict, Any, Optional
import uvicorn
from pydantic import BaseModel
from dotenv import load_dotenv
import openai
import requests
import re
import io
import pandas as pd
from fastapi.responses import Response
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Alignment
import glob
import uuid
import shutil

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("python_server_log.txt"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("proofmate")

# Load environment variables
load_dotenv()

# Set OpenAI API key and base URL
api_key = os.getenv("OPENAI_API_KEY")
api_base = os.getenv("OPENAI_API_BASE")

# Configure OpenAI client
openai.api_key = api_key
openai.base_url = api_base

# Get other environment variables
node_server_url = os.getenv("NODE_SERVER_URL", "http://localhost:5000")
environment = os.getenv("ENVIRONMENT", "development")

logger.info(f"OpenAI API Key: {api_key[:5]}...{api_key[-5:] if api_key else None}")
logger.info(f"OpenAI Base URL: {api_base}")
logger.info(f"Environment: {environment}")
logger.info(f"Node Server URL: {node_server_url}")

app = FastAPI(title="ProofMate - Notebook Analysis API")

# Configure CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Models
class ErrorHighlight(BaseModel):
    cell_index: int
    line_start: int
    line_end: int
    error_type: str
    error_message: str
    suggestion: str

class AnalysisResult(BaseModel):
    error_summary: str
    detailed_feedback: Dict[str, Any]
    confidence_score: float
    grade: float
    cell_annotations: List[Dict[str, Any]]
    error_highlights: List[ErrorHighlight] = []

# Utility functions
def extract_cells_from_notebook(notebook_content):
    """Parse notebook and extract cells with their content and metadata."""
    try:
        nb = nbformat.reads(notebook_content.decode('utf-8'), as_version=4)
        cells = []
        
        for i, cell in enumerate(nb.cells):
            if cell.cell_type == 'code':
                cells.append({
                    'index': i,
                    'type': 'code',
                    'content': cell.source,
                    'outputs': cell.outputs if hasattr(cell, 'outputs') else []
                })
            elif cell.cell_type == 'markdown':
                cells.append({
                    'index': i,
                    'type': 'markdown',
                    'content': cell.source
                })
        
        return cells
    except Exception as e:
        logger.error(f"Error parsing notebook: {str(e)}")
        return None

def detect_math_topic(cells):
    """Detect the mathematical topic from notebook cells."""
    full_content = " ".join([cell.get('content', '') for cell in cells])
    
    topics = {
        'linear_algebra': ['matrix', 'vector', 'eigenvalue', 'eigenvector', 'determinant', 'linear system'],
        'calculus': ['derivative', 'integral', 'limit', 'differential', 'integration'],
        'geometry': ['ellipse', 'circle', 'parabola', 'hyperbola', 'conic section'],
        'statistics': ['probability', 'distribution', 'mean', 'variance', 'regression'],
        'number_theory': ['prime', 'divisor', 'modulo', 'congruence', 'diophantine']
    }
    
    topic_counts = {}
    for topic, keywords in topics.items():
        count = sum(1 for keyword in keywords if keyword.lower() in full_content.lower())
        topic_counts[topic] = count
    
    if all(count == 0 for count in topic_counts.values()):
        return "general_mathematics"
    
    return max(topic_counts.items(), key=lambda x: x[1])[0]

def create_prompt_for_analysis(topic, reference_nb_repr, student_nb_repr):
    """Create a specialized prompt based on the detected mathematical topic."""
    
    topic_specific_instructions = {
        'linear_algebra': "Обрати внимание на операции с матрицами, векторные пространства, собственные значения/векторы и линейные преобразования.",
        'calculus': "Обрати внимание на вычисление производных, техники интегрирования, вычисление пределов и их применение.",
        'geometry': "Обрати внимание на конические сечения, координатную геометрию, преобразования и геометрические построения.",
        'statistics': "Обрати внимание на анализ данных, вероятностные расчеты, проверку гипотез и статистическое моделирование.",
        'number_theory': "Обрати внимание на простые числа, делимость, модульную арифметику и алгебраические структуры.",
        'general_mathematics': "Обрати внимание на правильность вычислений, математические рассуждения и реализацию алгоритмов."
    }
    
    return f"""
    Ты профессиональный математик и преподаватель, который анализирует работу студента.
    {topic_specific_instructions.get(topic, topic_specific_instructions['general_mathematics'])}
    
    # Эталонное решение:
    ```python
    {reference_nb_repr}
    ```
    
    # Решение студента:
    ```python
    {student_nb_repr}
    ```
    
    Проанализируй решение студента по сравнению с эталонным решением и предоставь детальный анализ по следующей структуре на русском языке:

    ## Краткое резюме
    [Дай краткое и конкретное резюме об общем качестве решения и основных проблемах]

    ## Сильные стороны
    [Перечисли 3-5 конкретных сильных сторон решения, каждый пункт должен быть уникальным]
    - Пункт 1
    - Пункт 2
    - и т.д.

    ## Области для улучшения
    [Перечисли 3-5 конкретных слабых сторон или ошибок, каждый пункт должен быть уникальным]
    - Пункт 1
    - Пункт 2
    - и т.д.

    ## Рекомендации
    [Перечисли 3-5 конкретных предложений по улучшению, каждый пункт должен быть уникальным]
    - Пункт 1
    - Пункт 2
    - и т.д.

    ## Комментарии к ячейкам
    [Для каждой ячейки с проблемами дай конкретный комментарий о проблеме и как её решить]
    Ячейка X: [конкретный комментарий для ячейки X]
    Ячейка Y: [конкретный комментарий для ячейки Y]
    и т.д.

    ## Оценка и уверенность
    Оценка: [Поставь оценку от 0 до 10, где 10 - идеальное решение]
    Уверенность: [Укажи уровень уверенности от 0 до 1, где 1 - полная уверенность]

    ВАЖНО: 
    1. Каждый пункт в разделах "Сильные стороны", "Области для улучшения" и "Рекомендации" должен быть уникальным - НЕ ПОВТОРЯЙ одну и ту же мысль разными словами.
    2. Обязательно используй формат списков с тире (-) для всех перечислений.
    3. Давай конкретные и полезные комментарии для каждой проблемной ячейки.
    4. Отвечай ТОЛЬКО на русском языке.
    """

def parse_ai_response(response_text):
    """Parse the AI response into structured feedback."""
    logger.info("Parsing AI response")
    
    # Extract summary - use the first paragraph that's not empty
    paragraphs = [p.strip() for p in response_text.split('\n\n') if p.strip()]
    error_summary = paragraphs[0] if paragraphs else "Анализ завершен."
    
    if len(error_summary) < 20 and len(paragraphs) > 1:
        # First paragraph might be too short, try the next one
        error_summary = paragraphs[1]
    
    # Try to extract grade - support both English and Russian patterns
    grade_match = re.search(r'(?:grade|оценка):?\s*(\d+(?:\.\d+)?)', response_text, re.IGNORECASE)
    grade = float(grade_match.group(1)) if grade_match else 7.5  # Default grade
    
    # Try to extract confidence - support both English and Russian patterns
    confidence_match = re.search(r'(?:confidence|уверенность):?\s*(\d+(?:\.\d+)?)', response_text, re.IGNORECASE)
    confidence = float(confidence_match.group(1)) if confidence_match else 0.9  # Default confidence
    
    # Extract strengths - use unified set of patterns for Russian and English
    strengths = []
    strength_section = re.search(r'##\s*(?:Strengths|Сильные\s+стороны)(.*?)(?=##|$)', response_text, re.IGNORECASE | re.DOTALL)
    
    if strength_section:
        # Extract all bullet points inside the strengths section
        bullet_points = re.findall(r'[-*•]\s*(.*?)(?=\n[-*•]|\n\n|$)', strength_section.group(1), re.DOTALL)
        strengths.extend([s.strip() for s in bullet_points if s.strip()])
    
        # If we didn't find strengths in a dedicated section, try direct text patterns
    if not strengths:
        # Try multiple patterns for strength extraction
        strength_patterns = [
            r'(?:strength|сильн[а-я]+\s+сторон[а-я]+|положительн[а-я]+)(?:[s:]\s*|\s*:\s*)(.*?)(?=\n|$)',
            r'(?:strength|сильн[а-я]+\s+сторон[а-я]+)[^\n:]*\n\s*[-*•]?\s*(.*?)(?=\n|$)',
            r'(?<=\n)[-*•]\s*(.*?)(?=\n|$)'  # Look for bullet points after strength headers
        ]
        
        for pattern in strength_patterns:
            matched_strengths = re.findall(pattern, response_text, re.IGNORECASE | re.MULTILINE)
            if matched_strengths:
                strengths.extend([s.strip() for s in matched_strengths if s.strip()])
    
    # Remove duplicates while preserving order
    unique_strengths = []
    seen_strengths = set()
    for s in strengths:
        normalized = s.lower().strip()
        if normalized not in seen_strengths and len(normalized) > 5:  # Only consider substantial items
            unique_strengths.append(s)
            seen_strengths.add(normalized)
    
    strengths = unique_strengths
    
    # Extract weaknesses with similar approach
    weaknesses = []
    weakness_section = re.search(r'##\s*(?:(?:Areas\s+for\s+Improvement|Weaknesses)|(?:Области\s+для\s+улучшения|Недостатки))(.*?)(?=##|$)', response_text, re.IGNORECASE | re.DOTALL)
    
    if weakness_section:
        # Extract all bullet points inside the weaknesses section
        bullet_points = re.findall(r'[-*•]\s*(.*?)(?=\n[-*•]|\n\n|$)', weakness_section.group(1), re.DOTALL)
        weaknesses.extend([w.strip() for w in bullet_points if w.strip()])
    
    # If we didn't find weaknesses in a dedicated section, try direct text patterns
    if not weaknesses:
        weakness_patterns = [
            r'(?:weakness|issue|error|problem|област[а-я]+\s+для\s+улучшени[а-я]+|недостат[а-я]+|ошибк[а-я]+)(?:[s:]\s*|\s*:\s*)(.*?)(?=\n|$)',
            r'(?:weakness|issue|error|problem|област[а-я]+\s+для\s+улучшени[а-я]+)[^\n:]*\n\s*[-*•]?\s*(.*?)(?=\n|$)'
        ]
        
        for pattern in weakness_patterns:
            matched_weaknesses = re.findall(pattern, response_text, re.IGNORECASE | re.MULTILINE)
            if matched_weaknesses:
                weaknesses.extend([w.strip() for w in matched_weaknesses if w.strip()])
    
    # Remove duplicates while preserving order
    unique_weaknesses = []
    seen_weaknesses = set()
    for w in weaknesses:
        normalized = w.lower().strip()
        if normalized not in seen_weaknesses and len(normalized) > 5:  # Only consider substantial items
            unique_weaknesses.append(w)
            seen_weaknesses.add(normalized)
    
    weaknesses = unique_weaknesses
    
    # Extract suggestions with similar approach
    suggestions = []
    suggestion_section = re.search(r'##\s*(?:Recommendations|Рекомендации)(.*?)(?=##|$)', response_text, re.IGNORECASE | re.DOTALL)
    
    if suggestion_section:
        # Extract all bullet points inside the recommendations section
        bullet_points = re.findall(r'[-*•]\s*(.*?)(?=\n[-*•]|\n\n|$)', suggestion_section.group(1), re.DOTALL)
        suggestions.extend([s.strip() for s in bullet_points if s.strip()])
    
    # If we didn't find suggestions in a dedicated section, try direct text patterns
    if not suggestions:
        suggestion_patterns = [
            r'(?:suggestion|recommendation|improvement|рекомендаци[а-я]+)(?:[s:]\s*|\s*:\s*)(.*?)(?=\n|$)',
            r'(?:suggestion|recommendation|improvement|рекомендаци[а-я]+)[^\n:]*\n\s*[-*•]?\s*(.*?)(?=\n|$)'
        ]
        
        for pattern in suggestion_patterns:
            matched_suggestions = re.findall(pattern, response_text, re.IGNORECASE | re.MULTILINE)
            if matched_suggestions:
                suggestions.extend([s.strip() for s in matched_suggestions if s.strip()])
    
    # Remove duplicates while preserving order
    unique_suggestions = []
    seen_suggestions = set()
    for s in suggestions:
        normalized = s.lower().strip()
        if normalized not in seen_suggestions and len(normalized) > 5:  # Only consider substantial items
            unique_suggestions.append(s)
            seen_suggestions.add(normalized)
    
    suggestions = unique_suggestions
    
    # Extract cell annotations
    cell_annotations = []
    
    # Look for a Cell Annotations section in the response
    cell_section_match = re.search(r'(?:##?\s*(?:Cell\s+Annotations|Комментарии\s+к\s+ячейкам))(.*?)(?=##|\Z)', 
                                 response_text, re.IGNORECASE | re.DOTALL)
    
    if cell_section_match:
        cell_section = cell_section_match.group(1).strip()
        
        # Extract annotations from bullet points with cell references
        # This pattern looks for: Ячейка X: or Cell X:
        bullet_cell_annotations = re.findall(
            r'(?:cell|ячейка)\s*(\d+)[:\s-]+\s*(.*?)(?=\n\s*(?:cell|ячейка)|\n\n|\Z)', 
            cell_section, 
            re.IGNORECASE | re.DOTALL
        )
        
        for cell_idx, comment in bullet_cell_annotations:
            try:
                cell_index = int(cell_idx.strip())
                comment_text = comment.strip()
                
                # Check if this cell already has annotations
                existing_cell = next((c for c in cell_annotations if c["cell_index"] == cell_index), None)
                if existing_cell:
                    existing_cell["comments"].append(comment_text)
                else:
                    cell_annotations.append({
                        "cell_index": cell_index,
                        "comments": [comment_text]
                    })
            except ValueError:
                logger.warning(f"Could not parse cell index from: {cell_idx}")
    
    # If we didn't find cell annotations in a dedicated section, try other patterns
    if not cell_annotations:
        # Look for specific cell mentions throughout the text - both English and Russian
        cell_mention_patterns = [
            r'(?:cell|ячейка|код\s+в\s+ячейке)\s*(\d+).*?[:：](.*?)(?=\n\s*(?:cell|ячейка|\n|$))',
            r'(?:cell|ячейка|код\s+в\s+ячейке)\s*(\d+)[^\n:]*\n\s*[-*•]?\s*(.*?)(?=\n|$)'
        ]
        
        for pattern in cell_mention_patterns:
            cell_mentions = re.findall(pattern, response_text, re.IGNORECASE | re.DOTALL)
            for cell_idx, comment in cell_mentions:
                try:
                    cell_index = int(cell_idx.strip())
                    comment_text = comment.strip()
                    
                    # Check if this cell already has annotations
                    existing_cell = next((c for c in cell_annotations if c["cell_index"] == cell_index), None)
                    if existing_cell:
                        existing_cell["comments"].append(comment_text)
                    else:
                        cell_annotations.append({
                            "cell_index": cell_index,
                            "comments": [comment_text]
                        })
                except ValueError:
                    logger.warning(f"Could not parse cell index from: {cell_idx}")
    
    # If we still don't have meaningful data, extract it from structured sections
    # This fallback makes the function more robust
    if not strengths and not weaknesses and len(cell_annotations) < 2:
        logger.warning("Regular parsing patterns didn't extract enough information. Trying fallback extraction methods.")
        
        # Look for structured sections in the response
        sections = {}
        current_section = None
        
        for line in response_text.split('\n'):
            line = line.strip()
            if not line:
                continue
                
            # Check if this line is a section header
            if re.match(r'^#+\s+\w+|^[A-ZА-Я][A-ZА-Яa-zа-я\s]+:', line):
                current_section = line.split(':', 1)[0].strip('# ').lower()
                sections[current_section] = []
            elif current_section and line.startswith('-') or line.startswith('*'):
                sections[current_section].append(line[1:].strip())
        
        # Extract data from identified sections
        for section_name, items in sections.items():
            if any(keyword in section_name for keyword in ['strength', 'сильн', 'положительн']):
                strengths.extend(items)
            elif any(keyword in section_name for keyword in ['weakness', 'issue', 'error', 'problem', 'област', 'недостат', 'ошибк']):
                weaknesses.extend(items)
            elif any(keyword in section_name for keyword in ['suggestion', 'recommendation', 'рекомендаци']):
                suggestions.extend(items)
    
    # Add default values if we still don't have anything
    if not strengths:
        strengths = ["Решение демонстрирует понимание основных математических концепций"]
    if not weaknesses and ("error" in error_summary.lower() or "ошибк" in error_summary.lower()):
        # Extract weakness from the summary if possible
        weaknesses = [error_summary]
    
    # Build the structured feedback
    detailed_feedback = {
        "strengths": strengths[:5],  # Limit to top 5 strengths
        "weaknesses": weaknesses[:5],  # Limit to top 5 weaknesses
        "suggestions": suggestions[:5] if suggestions else ["Ознакомьтесь с комментариями к ячейкам для детальных рекомендаций"]
    }
    
    # Log the extracted information for debugging
    logger.info(f"Extracted {len(strengths)} strengths, {len(weaknesses)} weaknesses, {len(suggestions)} suggestions, and {len(cell_annotations)} cell annotations")
    
    return {
        "error_summary": error_summary,
        "detailed_feedback": detailed_feedback,
        "confidence_score": min(max(confidence, 0), 1),  # Ensure between 0 and 1
        "grade": min(max(grade, 0), 10),  # Ensure between 0 and 10
        "cell_annotations": cell_annotations
    }

def call_openai_api_alternative(prompt, model="gpt-4o"):
    """
    Alternative method to call OpenAI API using requests library directly.
    """
    api_key = os.getenv("OPENAI_API_KEY")
    base_url = os.getenv("OPENAI_API_BASE")
    
    # Ensure base URL ends with /v1
    if base_url and not base_url.endswith('/v1'):
        base_url = f"{base_url}/v1"
    
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": "You are an AI assistant that analyzes Jupyter notebooks for mathematical problems."},
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.3
    }
    
    try:
        logger.info(f"Using alternative method to call OpenAI API with model: {model}")
        # Fix: Make sure to add a slash between base URL and chat/completions
        api_endpoint = f"{base_url}/chat/completions"
        logger.info(f"API endpoint: {api_endpoint}")
        
        response = requests.post(
            api_endpoint, 
            headers=headers,
            json=payload,
            timeout=30
        )
        
        if response.status_code == 200:
            result = response.json()
            content = result["choices"][0]["message"]["content"]
            logger.info(f"✅ Alternative method successful")
            return content
        else:
            logger.error(f"❌ Alternative method failed with status {response.status_code}: {response.text}")
            return None
    except Exception as e:
        logger.error(f"❌ Alternative method exception: {str(e)}")
        return None

# Utility function to create Excel report from analysis results
def create_excel_report(task_id: str, submissions_data: List[Dict[str, Any]]):
    """
    Создание компактного Excel-отчета из результатов анализа
    
    Аргументы:
        task_id: ID задания
        submissions_data: Список решений с результатами анализа
        
    Возвращает:
        BytesIO объект, содержащий Excel файл
    """
    logger.info(f"Создание Excel-отчета для задания {task_id} с {len(submissions_data)} решениями")
    
    # Создаем BytesIO объект для хранения Excel файла
    output = io.BytesIO()
    
    try:
        # Создаем Pandas Excel writer с использованием BytesIO объекта
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Создаем основной лист сводки - компактная версия
            summary_data = {
                "ID студента": [],
                "Имя": [],
                "Оценка": [],
                "Уверенность": [],
                "Дата сдачи": [],
                "Количество ошибок": [],
                "Комментарий": []
            }
            
            # Извлекаем данные из каждого решения
            for sub in submissions_data:
                analysis_result = sub.get("analysis_result", {})
                error_highlights = analysis_result.get("error_highlights", [])
                weaknesses = analysis_result.get("detailed_feedback", {}).get("weaknesses", [])
                
                # Вычисляем количество ошибок - используем либо error_highlights, либо количество слабых сторон
                error_count = len(error_highlights) if error_highlights else len(weaknesses)
                
                # Получаем краткую сводку об ошибках
                error_summary = analysis_result.get("error_summary", "Сводка об ошибках недоступна")
                
                # Добавляем данные сводки
                summary_data["ID студента"].append(sub.get("student_id", "Неизвестно"))
                summary_data["Имя"].append(sub.get("name", "Неизвестно"))
                summary_data["Оценка"].append(analysis_result.get("grade", 0.0))
                summary_data["Уверенность"].append(analysis_result.get("confidence_score", 0.0))
                summary_data["Дата сдачи"].append(sub.get("submission_date", ""))
                summary_data["Количество ошибок"].append(error_count)
                summary_data["Комментарий"].append(error_summary)
            
            # Преобразуем в DataFrame
            summary_df = pd.DataFrame(summary_data)
            
            # Записываем сводку в Excel лист
            summary_df.to_excel(writer, sheet_name="Сводка", index=False)
            
            # Форматируем лист сводки
            worksheet = writer.sheets["Сводка"]
            
            # Делаем заголовки жирными
            for col_num, value in enumerate(summary_df.columns.values):
                cell = worksheet.cell(row=1, column=col_num+1)
                cell.style = 'Headline 1'
                
                # Регулируем ширину столбцов для удобочитаемости
                if value == "Комментарий":
                    worksheet.column_dimensions[chr(65 + col_num)].width = 60  # Делаем столбец комментариев шире
                else:
                    worksheet.column_dimensions[chr(65 + col_num)].width = 15
            
            # Применяем перенос текста для комментариев
            for row in range(2, len(summary_data["Комментарий"]) + 2):
                cell = worksheet.cell(row=row, column=7)  # Столбец с комментариями
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                worksheet.row_dimensions[row].height = 60  # Регулируем высоту строки
            
            # Удалили создание листа с детальной обратной связью, чтобы отчет был компактным
            # и сфокусированным на информации о нескольких студенческих решениях
        
        # Получаем содержимое Excel файла
        output.seek(0)
        return output
    
    except Exception as e:
        logger.error(f"Error creating Excel file: {str(e)}")
        raise e

# Routes
@app.get("/")
async def root():
    return {"message": "API анализа ноутбуков ProofMate запущено"}

@app.get("/healthcheck")
async def healthcheck():
    return {"status": "ок", "environment": environment}

# Utility function to ensure analysis files are available
def ensure_analysis_files_available():
    """Copy analysis files from parent directory if they exist there but not in current directory."""
    current_dir = os.getcwd()
    parent_dir = os.path.abspath(os.path.join(current_dir, ".."))
    
    for filename in ["direct_analysis_result.json", "analysis_result.json"]:
        parent_path = os.path.join(parent_dir, filename)
        current_path = os.path.join(current_dir, filename)
        
        if os.path.exists(parent_path) and not os.path.exists(current_path):
            try:
                shutil.copy(parent_path, current_path)
                logger.info(f"Copied {filename} from parent directory to current directory")
            except Exception as e:
                logger.error(f"Failed to copy {filename}: {str(e)}")

@app.post("/api/analyze", response_model=AnalysisResult)
async def analyze_notebook(
    notebook_file: UploadFile = File(...),
    reference_solution: UploadFile = File(...),
    task_id: str = Form(...),
    student_id: str = Form(None),
    student_name: str = Form(None)
):
    """
    Analyze a student's notebook against a reference solution.
    Returns detailed feedback, error analysis, and a grade.
    """
    logger.info(f"Received analysis request for task {task_id}")
    logger.info(f"Student notebook: {notebook_file.filename}, Reference: {reference_solution.filename}")
    
    # Use filename as student name if not provided
    if not student_name:
        student_name = notebook_file.filename.split('.')[0] if notebook_file.filename else "Анонимный"
    
    # Generate a student ID based on name if not provided
    if not student_id:
        # Create a consistent ID based on student name and task ID
        # This ensures the same student gets the same ID for the same task
        # even if they submit multiple times
        name_for_id = student_name.lower().replace(" ", "_")
        student_id = f"{name_for_id}_{task_id}"[:8]
        
        # If the ID is too short or empty, add a random suffix
        if len(student_id) < 4:
            student_id += str(uuid.uuid4())[:4]
    
    logger.info(f"Processing submission for student ID: {student_id}, name: {student_name}")
    
    try:
        # Read the contents of both files
        student_content = await notebook_file.read()
        reference_content = await reference_solution.read()
        
        # Validate the file content
        if len(student_content) < 10 or len(reference_content) < 10:
            logger.error(f"One or both files appear to be empty or too small")
            raise HTTPException(status_code=400, detail="Один или оба файла ноутбуков пусты или недействительны")
        
        # Parse notebooks
        student_cells = extract_cells_from_notebook(student_content)
        reference_cells = extract_cells_from_notebook(reference_content)
        
        if not student_cells or not reference_cells:
            logger.error("Failed to parse notebook files")
            raise HTTPException(status_code=400, detail="Не удалось проанализировать файлы ноутбуков. Убедитесь, что это допустимые Jupyter notebooks.")
        
        # Detect the mathematical topic
        topic = detect_math_topic(student_cells + reference_cells)
        logger.info(f"Detected mathematical topic: {topic}")
        
        # Create simplified representations for the API call
        student_nb_repr = student_content.decode('utf-8')
        reference_nb_repr = reference_content.decode('utf-8')
        
        # Truncate if too large
        max_chars = 15000  # Reasonable limit for API
        if len(student_nb_repr) > max_chars:
            logger.info(f"Student notebook content too large ({len(student_nb_repr)} chars), truncating")
            student_nb_repr = student_nb_repr[:max_chars] + "... [truncated]"
        if len(reference_nb_repr) > max_chars:
            logger.info(f"Reference notebook content too large ({len(reference_nb_repr)} chars), truncating")
            reference_nb_repr = reference_nb_repr[:max_chars] + "... [truncated]"
        
        # Create analysis prompt
        analysis_prompt = create_prompt_for_analysis(topic, reference_nb_repr, student_nb_repr)
        
        # Use the direct HTTP method as it's known to work
        logger.info("Calling OpenAI API using direct HTTP request...")
        ai_response = call_openai_api_alternative(analysis_prompt)
        
        if not ai_response:
            # If direct method fails, try the SDK as fallback
            logger.info("Direct HTTP request failed, trying SDK as fallback...")
            try:
                response = openai.chat.completions.create(
                    model="gpt-4o",
                    messages=[
                        {"role": "system", "content": "Вы ИИ-ассистент, который анализирует математические решения."},
                        {"role": "user", "content": analysis_prompt}
                    ],
                    temperature=0.5,
                    max_tokens=4000
                )
                
                ai_response = response.choices[0].message.content
                logger.info("OpenAI API call successful via SDK")
            except Exception as e:
                logger.error(f"Both API call methods failed: {str(e)}")
                raise HTTPException(
                    status_code=500, 
                    detail="Не удалось получить ответ от API OpenAI. Пожалуйста, попробуйте позже."
                )
        
        # Log first part of the response for debugging
        logger.info(f"AI response preview: {ai_response[:200]}...")
        
        # Parse the AI response
        analysis_result = parse_ai_response(ai_response)
        
        # Save the raw response for debugging
        with open(f"response_debug_{task_id}_{student_id}.txt", "w") as f:
            f.write(ai_response)
        
        # Create directories for submissions if they don't exist
        submissions_dir = os.path.join(os.getcwd(), "submissions")
        if not os.path.exists(submissions_dir):
            os.makedirs(submissions_dir)
        
        task_dir = os.path.join(submissions_dir, task_id)
        if not os.path.exists(task_dir):
            os.makedirs(task_dir)
        
        student_dir = os.path.join(task_dir, student_id)
        if not os.path.exists(student_dir):
            os.makedirs(student_dir)
        
        # Save the analysis result for this student
        submission_info = {
            "student_id": student_id,
            "name": student_name,
            "email": "",  # Could add email field in future
            "submission_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "analysis_result": analysis_result
        }
        
        # Save to student's directory
        student_result_path = os.path.join(student_dir, "analysis_result.json")
        with open(student_result_path, "w", encoding='utf-8') as f:
            json.dump(submission_info, f, indent=2)
            logger.info(f"Saved analysis result for student {student_id} to: {student_result_path}")
        
        # Also save to standard locations for backward compatibility
        with open("analysis_result.json", "w", encoding='utf-8') as f:
            json.dump(analysis_result, f, indent=2)
            logger.info(f"Saved analysis result to current directory")
        
        # Also save to parent directory for access from web client
        try:
            parent_dir = os.path.abspath(os.path.join(os.getcwd(), ".."))
            parent_file_path = os.path.join(parent_dir, "analysis_result.json")
            with open(parent_file_path, "w", encoding='utf-8') as f:
                json.dump(analysis_result, f, indent=2)
                logger.info(f"Saved analysis result to parent directory: {parent_file_path}")
        except Exception as e:
            logger.error(f"Failed to save analysis result to parent directory: {str(e)}")
        
        # Validate the analysis result
        if not analysis_result["detailed_feedback"]["strengths"] and not analysis_result["detailed_feedback"]["weaknesses"]:
            logger.warning("Analysis result lacks both strengths and weaknesses")
            # Add a default strength if none were extracted
            analysis_result["detailed_feedback"]["strengths"] = ["Решение демонстрирует понимание основных математических концепций"]
        
        if not analysis_result["cell_annotations"]:
            logger.warning("Analysis result lacks cell annotations")
            # Try to extract some cell-level information from the weaknesses
            for weakness in analysis_result["detailed_feedback"]["weaknesses"]:
                cell_match = re.search(r'(?:cell|ячейка)\s*(\d+)', weakness, re.IGNORECASE)
                if cell_match:
                    cell_index = int(cell_match.group(1))
                    analysis_result["cell_annotations"].append({
                        "cell_index": cell_index,
                        "comments": [weakness]
                    })
        
        # Return the analysis result
        return AnalysisResult(
            error_summary=analysis_result["error_summary"],
            detailed_feedback=analysis_result["detailed_feedback"],
            confidence_score=analysis_result["confidence_score"],
            grade=analysis_result["grade"],
            cell_annotations=analysis_result["cell_annotations"],
            error_highlights=[]  # Empty for now, will be enhanced in future
        )
        
    except Exception as e:
        logger.error(f"Error analyzing notebooks: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Ошибка анализа: {str(e)}")

@app.get("/api/export-report/{task_id}")
async def export_report(task_id: str):
    """
    Генерация Excel-отчета для заданий конкретной задачи
    """
    logger.info(f"Generating Excel report for task ID: {task_id}")
    
    try:
        # Define the path to the submissions directory for this task
        submissions_dir = os.path.join(os.getcwd(), "submissions", task_id)
        
        # Initialize submissions data
        submissions_data = []
        
        # Check if the submissions directory exists
        if os.path.exists(submissions_dir) and os.path.isdir(submissions_dir):
            # Get all student subdirectories
            student_dirs = [d for d in os.listdir(submissions_dir) 
                           if os.path.isdir(os.path.join(submissions_dir, d))]
            
            logger.info(f"Found {len(student_dirs)} student submissions for task {task_id}")
            
            # Process each student's submission
            for student_id in student_dirs:
                student_dir = os.path.join(submissions_dir, student_id)
                analysis_file = os.path.join(student_dir, "analysis_result.json")
                
                if os.path.exists(analysis_file):
                    try:
                        with open(analysis_file, "r", encoding='utf-8') as f:
                            student_data = json.load(f)
                        
                        # Add to submissions data - include all submissions regardless of name
                        submissions_data.append(student_data)
                        logger.info(f"Added student {student_id} to the report")
                    except Exception as e:
                        logger.error(f"Error reading analysis result for student {student_id}: {str(e)}")
                else:
                    logger.warning(f"No analysis result found for student {student_id}")
        
        # If no submissions found in the new structure, try the legacy approach
        if not submissions_data:
            logger.info("Решения не найдены в новой структуре директорий. Пробуем устаревший подход.")
            
            # Make sure analysis files are available
            ensure_analysis_files_available()
            
            # Log current working directory for debugging
            current_dir = os.getcwd()
            parent_dir = os.path.abspath(os.path.join(current_dir, ".."))
            
            # Define list of possible file paths to check in various locations
            possible_paths = []
            
            # Look in current directory and parent directory for both files
            for filename in ["analysis_result.json", "direct_analysis_result.json"]:
                # Current directory
                possible_paths.append(os.path.join(current_dir, filename))
                # Parent directory 
                possible_paths.append(os.path.join(parent_dir, filename))
            
            # Find all existing files and pick the most recent one
            existing_files = []
            for file_path in possible_paths:
                if os.path.exists(file_path):
                    mod_time = os.path.getmtime(file_path)
                    file_size = os.path.getsize(file_path)
                    existing_files.append((file_path, mod_time, file_size))
            
            # Sort files by modification time (most recent first)
            existing_files.sort(key=lambda x: x[1], reverse=True)
            
            analysis_result = None
            if existing_files:
                # Use the most recent file
                most_recent_file = existing_files[0][0]
                logger.info(f"Using most recent file: {most_recent_file}")
                try:
                    with open(most_recent_file, "r", encoding='utf-8') as f:
                        analysis_result = json.load(f)
                    logger.info(f"Successfully loaded analysis result from {most_recent_file}")
                except Exception as e:
                    logger.error(f"Failed to load analysis result from {most_recent_file}: {str(e)}")
                    
                    # Try other files if available
                    for file_path, _, _ in existing_files[1:]:
                        try:
                            logger.info(f"Trying alternate file: {file_path}")
                            with open(file_path, "r", encoding='utf-8') as f:
                                analysis_result = json.load(f)
                            logger.info(f"Successfully loaded analysis result from {file_path}")
                            break
                        except Exception as e2:
                            logger.error(f"Failed to load analysis result from {file_path}: {str(e2)}")
            
            # If analysis result found, create a mock submission
            if analysis_result:
                # Extract notebook info and student information
                assignment_topic = "Неизвестная тема"
                student_name = "Студент"
                
                # Try to extract assignment topic and student name from filenames
                notebook_files = []
                student_files = glob.glob(os.path.join(parent_dir, "решение_студента*.ipynb"))
                notebook_files.extend(student_files)
                student_files = glob.glob(os.path.join(current_dir, "решение_студента*.ipynb"))
                notebook_files.extend(student_files)
                
                if notebook_files:
                    # Extract student name from filename
                    filename = os.path.basename(notebook_files[0])
                    if "_ellipse" in filename:
                        assignment_topic = "Эллипс"
                    elif "_complex" in filename:
                        assignment_topic = "Комплексные числа"
                    
                    # Try to get a more specific student name if available
                    student_name_match = re.search(r'(?:решение_студента|student_solution)_([^\.]+)\.ipynb', filename)
                    if student_name_match:
                        student_name = student_name_match.group(1).strip()
                        if not student_name:
                            student_name = "Студент"
                
                # Create a submission entry with the loaded analysis result
                submissions_data.append({
                    "student_id": f"{student_name}_{task_id}",
                    "name": f"{student_name} - {assignment_topic}",
                    "email": "студент@example.edu",
                    "submission_date": datetime.now().strftime("%Y-%m-%d"),
                    "analysis_result": analysis_result
                })
                logger.info(f"Создана запись решения по устаревшему методу для студента: {student_name}")
        
        # If still no submissions, create a dummy entry
        if not submissions_data:
            logger.warning("Не найдены корректные результаты анализа. Создаём тестовый отчёт.")
            submissions_data = [
                {
                    "student_id": "Неизвестно",
                    "name": "Тестовый студент",
                    "email": "студент@example.edu",
                    "submission_date": datetime.now().strftime("%Y-%m-%d"),
                    "analysis_result": {
                        "error_summary": "Это тестовый отчет анализа. Реальный анализ не проводился.",
                        "detailed_feedback": {
                            "strengths": ["Это тестовое преимущество"],
                            "weaknesses": ["Это тестовый недостаток"],
                            "suggestions": ["Это тестовое предложение"]
                        },
                        "confidence_score": 0.5,
                        "grade": 0.0,
                        "cell_annotations": [
                            {"cell_index": 0, "comments": ["Тестовый комментарий"]}
                        ],
                        "error_highlights": []
                    }
                }
            ]
        
        # Use our utility function to create the Excel report
        output = create_excel_report(task_id, submissions_data)
        excel_data = output.getvalue()
        
        # Create a response with the Excel file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"proofmate_отчет_{task_id}_{timestamp}.xlsx"
        
        headers = {
            'Content-Disposition': f'attachment; filename="{filename}"',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
        
        logger.info(f"Excel-отчёт успешно создан для задания: {task_id} с {len(submissions_data)} решениями")
        return Response(content=excel_data, headers=headers)
        
    except Exception as e:
        logger.error(f"Error generating Excel report for task ID {task_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Ошибка создания отчета: {str(e)}")

if __name__ == "__main__":
    port = int(os.environ.get("PYTHON_SERVER_PORT", 8000))
    uvicorn.run("main_functional:app", host="0.0.0.0", port=port, reload=True) 